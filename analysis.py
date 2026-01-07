
import json
import os
import glob
from datetime import datetime
import re
import openpyxl
from openpyxl.styles import PatternFill
import csv

def consolidate_logs(output_dir):
    """
    Consolidates all log files in the specified directory into a single JSON file.
    """
    consolidated_results = {}
    log_files = glob.glob(os.path.join(output_dir, "results_*.json"))
    log_files = [f for f in log_files if not f.endswith("_analyzed.json")]

    # Load climate data to get context
    climate_json_path = 'data/logic/climate.json'
    try:
        with open(climate_json_path, 'r', encoding='utf-8') as f:
            climate_data = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError) as e:
        print(f"Error reading climate data: {e}")
        climate_data = {}

    for log_file in log_files:
        try:
            with open(log_file, 'r', encoding='utf-8') as f:
                log_data = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            continue

        experiment_info = log_data.get("experiment_info", {})
        llm_name = experiment_info.get("llm_name")
        if not llm_name:
            continue

        if llm_name not in consolidated_results:
            consolidated_results[llm_name] = {
                "executions": []
            }
        
        start_time_str = experiment_info.get("start_time")
        end_time_str = experiment_info.get("end_time")
        duration_str = "Unknown"

        if start_time_str and end_time_str:
            try:
                start_time = datetime.strptime(start_time_str, "%Y%m%d_%H%M%S")
                end_time = datetime.strptime(end_time_str, "%Y%m%d_%H%M%S")
                duration = end_time - start_time
                minutes = duration.total_seconds() / 60
                duration_str = f"{minutes:.2f} minutes"
            except ValueError:
                duration_str = "Invalid format"

        execution_data = {
            "execution_time": duration_str,
            "questions": []
        }

        for question in log_data.get("questions", []):
            question_id = question.get("question_id")
            steps = question.get("steps", [])

            llm_fallacies = {}
            ground_truth = None
            is_correct = False
            
            # First, find all successful lean4 codes
            successful_lean4_codes = {}
            for step in steps:
                step_name = step.get("step_name", "")
                if "_stage3_verification_attempt_" in step_name:
                    match = re.search(r"option_(\d+)_stage3_verification_attempt_", step_name)
                    if match:
                        option_num = match.group(1)
                        if step.get("result", {}).get("pass"):
                            successful_lean4_codes[option_num] = step.get("code_sent_to_verifier")

            # Then, build the question summary
            for step in steps:
                step_name = step.get("step_name", "")
                if "_final_comparison" in step_name:
                    match = re.search(r"option_(\d+)_final_comparison", step_name)
                    if match:
                        option_num = match.group(1)
                        fallacy = step.get("option_llm_fallacy")
                        verification_status = step.get("option_verification_status")
                        lean_pass = verification_status in ["lean_pass", "lean_pass_with_type_error"]
                        
                        if fallacy:
                            llm_fallacies[option_num] = {
                                "fallacy_name": fallacy,
                                "lean_pass": lean_pass,
                                "option_verification_status": verification_status,
                                "option_text_explanation": step.get("option_text_explanation"),
                                "lean4_code": successful_lean4_codes.get(option_num, "")
                            }
                elif step_name == "final_summary":
                    ground_truth = step.get("ground_truth")
                    is_correct = step.get("overall_is_correct", False)

            question_summary = {
                "question_id": question_id,
                "is_correct": is_correct,
                "llm_fallacy": llm_fallacies,
                "ground_truth": ground_truth,
                "context": climate_data.get(question_id, {}).get("context"),
            }

            execution_data["questions"].append(question_summary)
        
        consolidated_results[llm_name]["executions"].append(execution_data)

    output_filepath = os.path.join(output_dir, "consolidated_results.json")
    
    try:
        with open(output_filepath, 'w', encoding='utf-8') as f:
            json.dump(consolidated_results, f, indent=4, ensure_ascii=False)
        print(f"--- Consolidation complete. New JSON saved to {output_filepath} ---")
    except Exception as e:
        print(f"Error: Could not write to {output_filepath}. Reason: {e}")


def convert_to_excel(json_filepath, excel_filepath):
    """
    Converts the consolidated JSON results to an Excel file and a CSV file, with conditional formatting in Excel.
    """
    try:
        with open(json_filepath, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except FileNotFoundError:
        print(f"Error: JSON file not found at {json_filepath}")
        return

    # --- Excel Setup ---
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Experiment Results"
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # --- Data Processing ---
    all_rows = []
    NUM_OPTIONS = 3 # We are now dealing with N options, not attempts

    # Define the header in the desired order
    header = [
        "model_name", "question_id", "context", "ground_truth",
        "fallacy_option1", "verification_status1",
        "fallacy_option2", "verification_status2",
        "fallacy_option3", "verification_status3",
        "lean4_code1", "explanation1",
        "lean4_code2", "explanation2",
        "lean4_code3", "explanation3"
    ]
    all_rows.append(header)

    # Process data for rows
    if data:
        for model_name, model_data in data.items():
            for execution in model_data.get("executions", []):
                for q in execution.get("questions", []):
                    row_data = {
                        "model_name": model_name,
                        "question_id": q.get("question_id"),
                        "context": q.get("context"),
                        "ground_truth": q.get("ground_truth")
                    }
                    
                    llm_fallacies = q.get("llm_fallacy", {})
                    for i in range(1, NUM_OPTIONS + 1):
                        option_data = llm_fallacies.get(str(i))
                        if option_data:
                            row_data[f"fallacy_option{i}"] = option_data.get("fallacy_name")
                            row_data[f"verification_status{i}"] = option_data.get("option_verification_status")
                            row_data[f"lean4_code{i}"] = option_data.get("lean4_code")
                            row_data[f"explanation{i}"] = option_data.get("option_text_explanation")
                        else:
                            row_data[f"fallacy_option{i}"] = None
                            row_data[f"verification_status{i}"] = None
                            row_data[f"lean4_code{i}"] = None
                            row_data[f"explanation{i}"] = None
                    
                    final_row = [row_data.get(h) for h in header]
                    all_rows.append(final_row)

    # --- Write to Excel ---
    for row in all_rows:
        sheet.append(row)

    # --- Conditional Formatting ---
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")   # Green
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")     # Light Red

    header = all_rows[0]
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row), start=2):
        for col_idx, cell in enumerate(row, start=1):
            col_name = header[col_idx - 1]
            if col_name and cell.value is not None:
                # Highlight correct answers in green
                if "_is_correct" in col_name and cell.value is True:
                    cell.fill = green_fill
                # Highlight lean_pass status
                if "verification_status" in col_name:
                    if "lean_pass" in str(cell.value):
                        cell.fill = green_fill
                    elif "no_pass" in str(cell.value):
                        cell.fill = red_fill

    try:
        workbook.save(excel_filepath)
        print(f"--- Excel conversion complete. File saved to {excel_filepath} ---")
    except Exception as e:
        print(f"Error: Could not write to {excel_filepath}. Reason: {e}")

    # --- Write to CSV ---
    csv_filepath = os.path.splitext(excel_filepath)[0] + '.csv'
    try:
        with open(csv_filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerows(all_rows)
        print(f"--- CSV conversion complete. File saved to {csv_filepath} ---")
    except Exception as e:
        print(f"Error: Could not write to {csv_filepath}. Reason: {e}")
